#!/usr/bin/env python3
"""
AMRWatch database seeder.
Reads actual mockdata/ Excel files as templates, then generates a
realistic multi-month dataset that makes all dashboard charts meaningful.

Tables populated (in dependency order):
  sites → water_samples → isolates → amr_sequences
                                    → wgs_metrics
"""

import json
import random
import uuid
from datetime import date, timedelta

import openpyxl
import psycopg2
from psycopg2.extras import Json, execute_values

# ── DB connection (matches application.yaml) ──────────────────────────────────
DB = dict(host="localhost", port=5432, dbname="amr_dashboard",
          user="postgres", password="postgres")

# ── Reference data (from mockdata files + standard MDRO classifications) ───────
SITES_RAW = [
    # siteId, locationName, riverName, lat, lng
    ("A10", "Farm A Dispatch",          "Apies River", -25.747,  28.229),
    ("B26", "Farm B Pivot 1",           "Apies River", -25.750,  28.230),
    ("B27", "Tshwane North",            "Apies River", -25.752,  28.231),
    ("C01", "Hammanskraal Bridge",      "Apies River", -25.466,  28.345),
    ("C02", "Roodeplaat Dam Inlet",     "Apies River", -25.615,  28.371),
    ("D14", "Pretoria West Outlet",     "Apies River", -25.734,  28.179),
    ("D15", "Silverton Industrial",     "Apies River", -25.726,  28.300),
    ("E03", "Sunderland Ridge",         "Apies River", -25.676,  28.249),
    ("E04", "Wallmansthal Crossing",    "Apies River", -25.540,  28.312),
    ("F07", "Babelegi Catchment",       "Apies River", -25.440,  28.390),
    ("G11", "Temba Irrigation Point",   "Apies River", -25.389,  28.411),
    ("G12", "Rethabiseng Confluence",   "Apies River", -25.372,  28.432),
    ("H09", "Ekangala Upstream",        "Apies River", -25.350,  28.469),
    ("H10", "Ekangala Downstream",      "Apies River", -25.340,  28.483),
]

# Organisms from Binary_Information.xlsx — matched to clinical MDRO codes.
# Weights proportional to isolate counts in the mockdata (n=90 total).
# MDRO codes: ESBL/CRE/VRE/MDR/MDRO feed the incident-rate endpoint.
ORGANISMS = [
    ("Morganella morganii",          "MDR"),    # 12 isolates
    ("Citrobacter freundii",         "ESBL"),   # 10 isolates
    ("Stenotrophomonas maltophilia", "MDR"),    # 9 isolates
    ("Serratia fonticola",           "MDRO"),   # 9 isolates
    ("Proteus mirabilis",            "ESBL"),   # 9 isolates
    ("Pseudomonas aeruginosa",       "MDR"),    # 7 isolates
    ("Salmonella enterica",          "ESBL"),   # 7 isolates
    ("Klebsiella pneumoniae",        "CRE"),    # 6 isolates
    ("Enterococcus faecalis",        "VRE"),    # 6 isolates
    ("Escherichia coli",             "ESBL"),   # 5 isolates
    ("Acinetobacter baumannii",      "MDRO"),   # 5 isolates
    ("Enterobacter cloacae",         "ESBL"),   # 5 isolates
]

# Organism → (virulence_genes, source_context)
# Virulence genes and contexts from Binary_Information.xlsx
ORG_META = {
    "Morganella morganii":          ("hlyA, set1",          "River water (downstream)"),
    "Citrobacter freundii":         ("aggR, aatA",          "Agricultural runoff"),
    "Stenotrophomonas maltophilia": (None,                   "Sediment sample"),
    "Serratia fonticola":           (None,                   "Irrigation pivot point"),
    "Proteus mirabilis":            ("fim, mrkD",            "River water (upstream)"),
    "Pseudomonas aeruginosa":       ("exoS, exoT",           "Effluent discharge point"),
    "Salmonella enterica":          ("sfa, espA",            "Agricultural runoff"),
    "Klebsiella pneumoniae":        ("rmpA, iutA, iroN",     "Spinach at harvest"),
    "Enterococcus faecalis":        ("fim, aggR, papA",      "Treated wastewater"),
    "Escherichia coli":             ("papA, fyuA",           "River water (downstream)"),
    "Acinetobacter baumannii":      ("bap, ompA",            "Hospital effluent"),
    "Enterobacter cloacae":         ("irp2, fyuA",           "Raw wastewater"),
}

# Organism → binary typing profile (Intl1/2/3 = integron presence; TEM/SHV = beta-lac markers)
# Int3=True: 64 % of isolates in mockdata; TEM=True: 51 %; SHV=True: 48 %
ORG_BINARY = {
    "Morganella morganii":          {"Intl1": True,  "Intl2": True,  "Intl3": True,  "TEM": True,  "SHV": False},
    "Citrobacter freundii":         {"Intl1": True,  "Intl2": False, "Intl3": True,  "TEM": True,  "SHV": True },
    "Stenotrophomonas maltophilia": {"Intl1": False, "Intl2": False, "Intl3": True,  "TEM": False, "SHV": False},
    "Serratia fonticola":           {"Intl1": False, "Intl2": True,  "Intl3": True,  "TEM": False, "SHV": False},
    "Proteus mirabilis":            {"Intl1": True,  "Intl2": False, "Intl3": True,  "TEM": True,  "SHV": True },
    "Pseudomonas aeruginosa":       {"Intl1": False, "Intl2": True,  "Intl3": False, "TEM": False, "SHV": False},
    "Salmonella enterica":          {"Intl1": True,  "Intl2": True,  "Intl3": True,  "TEM": True,  "SHV": False},
    "Klebsiella pneumoniae":        {"Intl1": True,  "Intl2": False, "Intl3": True,  "TEM": True,  "SHV": True },
    "Enterococcus faecalis":        {"Intl1": True,  "Intl2": True,  "Intl3": False, "TEM": False, "SHV": False},
    "Escherichia coli":             {"Intl1": True,  "Intl2": True,  "Intl3": False, "TEM": True,  "SHV": False},
    "Acinetobacter baumannii":      {"Intl1": True,  "Intl2": False, "Intl3": True,  "TEM": False, "SHV": False},
    "Enterobacter cloacae":         {"Intl1": False, "Intl2": True,  "Intl3": True,  "TEM": True,  "SHV": True },
}

# AMR genes from AMRFinderPlus_Results.xlsx — exact symbols, classes, and subclasses.
# All 25 unique genes present in the mockdata, with representative identity values.
AMR_GENES = [
    # (gene_symbol, element_type, resistance_class, resistance_subclass, avg_identity, avg_coverage)
    ("ermC",         "AMR", "MACROLIDE",     "MACROLIDE",         87.3,  85.0),
    ("mcr-1",        "AMR", "COLISTIN",      "COLISTIN",          89.2,  92.0),
    ("tet(B)",       "AMR", "TETRACYCLINE",  "TETRACYCLINE",      89.3,  90.0),
    ("qnrS",         "AMR", "QUINOLONE",     "FLUOROQUINOLONE",   87.1,  65.0),
    ("catA",         "AMR", "PHENICOL",      "CHLORAMPHENICOL",   92.0,  95.0),
    ("dfrA",         "AMR", "TRIMETHOPRIM",  "TRIMETHOPRIM",      86.6,  80.0),
    ("qnrB",         "AMR", "QUINOLONE",     "FLUOROQUINOLONE",   89.3,  72.0),
    ("blaTEM",       "AMR", "BETA-LACTAM",   "BETA-LACTAM",       86.4,  88.0),
    ("aph(3')-Ia",   "AMR", "AMINOGLYCOSIDE","KANAMYCIN",         86.5,  90.0),
    ("aac(6')-Ib",   "AMR", "AMINOGLYCOSIDE","AMIKACIN",          91.6,  88.0),
    ("blaOXA",       "AMR", "BETA-LACTAM",   "CARBAPENEM",        91.0,  78.1),
    ("sul2",         "AMR", "SULFONAMIDE",   "SULFONAMIDE",       82.0,  51.9),
    ("bla",          "AMR", "BETA-LACTAM",   "BETA-LACTAM",       82.7,  56.2),
    ("blaCTX",       "AMR", "BETA-LACTAM",   "CEPHALOSPORIN",     85.5,  80.0),
]

# WGS templates from StarAMR_Metrics.xlsx
# SIR profile weights: Resistant 37.8 %, Intermediate 25.6 %, Susceptible 36.7 %
WGS_TEMPLATES = [
    ("PASS", "aph(3')-Ia",             "ciprofloxacin, ceftriaxone",                        "Resistant",     "IncFIA",        6652054, 467886),
    ("PASS", "aph(3')-Ia",             "cefotaxime, erythromycin, florfenicol",              "Resistant",     "IncFIC(FII)",   5406711,  94870),
    ("PASS", "blaTEM",                  "tetracycline, gentamicin, imipenem",                 "Intermediate",  "IncX3",         4086939, 557478),
    ("PASS", "mcr-1",                   "colistin",                                           "Resistant",     "IncP1",         5214567,  87421),
    ("PASS", "ermC, tet(B)",            "erythromycin, tetracycline",                         "Intermediate",  "IncHI2",        5017831, 156657),
    ("PASS", "qnrS, dfrA",             "ciprofloxacin, trimethoprim",                        "Susceptible",   "IncFIB(K)",     4988123,  98342),
    ("PASS", "catA, blaTEM",            "chloramphenicol, ampicillin",                        "Susceptible",   "IncFIB(AP001918)", 5125000, 120000),
    ("FAIL", "blaOXA-48, dfrA1, catA1","ertapenem, vancomycin, imipenem",                    "Susceptible",   "IncFIC(FII)",   4711190,  92098),
    ("PASS", "sul2, qnrB",             "sulfisoxazole, ciprofloxacin",                       "Intermediate",  "IncN",          5300000, 200000),
    ("FAIL", "blaCTX, aac(6')-Ib",     "ceftriaxone, amikacin",                              "Resistant",     "IncI1",         5800000,  50000),
]

ANALYSIS_TYPES = ["WGS", "Metagenomics", "WGS", "16S rRNA", "Metagenomics", "Shotgun"]

# ── Helper ────────────────────────────────────────────────────────────────────
def read_excel_rows(path, sheet=0):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            data.append(dict(zip(headers, row)))
    return data

def rand_date(start: date, end: date) -> date:
    return start + timedelta(days=random.randint(0, (end - start).days))

def env_noise(base, pct=0.15):
    return round(base * (1 + random.uniform(-pct, pct)), 2)

# ── Main seeding logic ────────────────────────────────────────────────────────
def seed():
    conn = psycopg2.connect(**DB)
    cur  = conn.cursor()

    # ── Truncate all tables (cascade handles FK order) ──────────────────────
    print("Clearing existing data...")
    cur.execute("""
        TRUNCATE amr_sequences, isolates, water_samples, sites
        RESTART IDENTITY CASCADE
    """)
    conn.commit()

    # ── 1. SITES ────────────────────────────────────────────────────────────
    print(f"Inserting {len(SITES_RAW)} sites...")
    execute_values(cur,
        "INSERT INTO sites (site_id, location_name, river_name, latitude, longitude) VALUES %s",
        SITES_RAW
    )
    conn.commit()

    # ── 2. WATER SAMPLES ─────────────────────────────────────────────────────
    # Generate 6 trips × 14 sites = 84 samples, spread Jan 2025 – Apr 2026
    # Plus additional 2024 samples for YoY comparisons
    # Monthly sampling programme — one trip per month per site.
    # Each tuple: (trip_id, collection_date)
    # 2024 and 2025 are complete years; 2026 is Jan–Apr (year to date).
    TRIPS = [
        # ── 2024 (full year) ───────────────────────────────────────────
        ("2024-T01", date(2024, 1,  9)),
        ("2024-T02", date(2024, 2, 13)),
        ("2024-T03", date(2024, 3, 12)),
        ("2024-T04", date(2024, 4,  9)),
        ("2024-T05", date(2024, 5, 14)),
        ("2024-T06", date(2024, 6, 11)),
        ("2024-T07", date(2024, 7,  9)),
        ("2024-T08", date(2024, 8, 13)),
        ("2024-T09", date(2024, 9, 10)),
        ("2024-T10", date(2024, 10, 8)),
        ("2024-T11", date(2024, 11,12)),
        ("2024-T12", date(2024, 12,10)),
        # ── 2025 (full year) ───────────────────────────────────────────
        ("2025-T01", date(2025, 1, 14)),
        ("2025-T02", date(2025, 2, 11)),
        ("2025-T03", date(2025, 3, 18)),
        ("2025-T04", date(2025, 4,  8)),
        ("2025-T05", date(2025, 5, 13)),
        ("2025-T06", date(2025, 6, 10)),
        ("2025-T07", date(2025, 7, 15)),
        ("2025-T08", date(2025, 8, 12)),
        ("2025-T09", date(2025, 9,  9)),
        ("2025-T10", date(2025, 10,14)),
        ("2025-T11", date(2025, 11,11)),
        ("2025-T12", date(2025, 12, 9)),
        # ── 2026 (Jan–Apr, year to date) ──────────────────────────────
        ("2026-T01", date(2026, 1, 13)),
        ("2026-T02", date(2026, 2, 10)),
        ("2026-T03", date(2026, 3, 10)),
        ("2026-T04", date(2026, 4,  8)),
    ]

    # Base env params per site (slightly different microclimates)
    SITE_ENV = {s[0]: {"temp": 18.5 + random.uniform(-2, 4), "ph": 7.2 + random.uniform(-0.5, 0.5),
                        "tds": 250 + random.uniform(-30, 80), "ec": 400 + random.uniform(-50, 100),
                        "do": 6.5 + random.uniform(-1.5, 1.5)} for s in SITES_RAW}

    samples_insert = []
    sample_ids = []  # (sample_id, site_id, trip_id, collection_date)

    system_user = uuid.UUID("00000000-0000-0000-0000-000000000001")

    for trip_id, trip_date in TRIPS:
        for site in SITES_RAW:
            site_id = site[0]
            env = SITE_ENV[site_id]
            samp_num = len(sample_ids) + 1
            sample_id = f"SAMP-{samp_num:04d}"
            sample_name = f"{site_id}-{trip_date.strftime('%b').upper()}{str(trip_date.year)[2:]}"

            samples_insert.append((
                sample_id, site_id, str(system_user), trip_id,
                trip_date,
                env_noise(env["temp"]),
                env_noise(env["ph"], 0.05),
                env_noise(env["tds"]),
                env_noise(env["ec"]),
                env_noise(env["do"]),
                sample_name,
                random.choice(ANALYSIS_TYPES),
            ))
            sample_ids.append((sample_id, site_id, trip_id, trip_date))

    print(f"Inserting {len(samples_insert)} water samples...")
    execute_values(cur, """
        INSERT INTO water_samples
          (sample_id, site_id, collected_by_user_id, trip_identifier,
           collection_date, water_temperature, ph_level, tds, ec,
           dissolved_oxygen, sample_name, sample_analysis_type)
        VALUES %s
    """, samples_insert)
    conn.commit()

    # ── 3. ISOLATES ──────────────────────────────────────────────────────────
    # 2–4 isolates per sample, weighted proportional to mockdata counts (n=90).
    # 12 organisms: weights sum to 100, mirroring Binary_Information.xlsx distribution.
    ORG_WEIGHTS = [12, 10, 9, 9, 9, 7, 7, 6, 6, 5, 5, 5]  # matches ORGANISMS list order
    isolates_insert = []
    isolate_ids = []   # (isolate_id, sample_id)
    iso_counter = 100

    for (sample_id, site_id, trip_id, trip_date) in sample_ids:
        n_isolates = random.choices([1, 2, 3, 4], weights=[10, 40, 35, 15])[0]
        chosen_orgs = random.choices(ORGANISMS, weights=ORG_WEIGHTS, k=n_isolates)

        for idx, (org_name, ar_code) in enumerate(chosen_orgs):
            isolate_id   = f"ISO-{iso_counter}"
            isolate_num  = f"{site_id}{idx + 1}H"
            vgenes, ctx  = ORG_META[org_name]
            binary_prof  = ORG_BINARY[org_name]
            iso_counter += 1

            isolates_insert.append((
                isolate_id, sample_id, str(system_user),
                isolate_num, org_name, ctx, ar_code,
                vgenes,
                Json(binary_prof),
            ))
            isolate_ids.append((isolate_id, org_name))

    print(f"Inserting {len(isolates_insert)} isolates...")
    execute_values(cur, """
        INSERT INTO isolates
          (isolate_id, sample_id, owner_id, isolate_number,
           organism_identity, source_context, ar_code, virulence_genes,
           binary_typing_profile)
        VALUES %s
    """, isolates_insert)
    conn.commit()

    # ── 4. AMR SEQUENCES ────────────────────────────────────────────────────
    # 1–3 genes per isolate
    amr_insert = []
    for (isolate_id, org_name) in isolate_ids:
        n_genes = random.choices([1, 2, 3], weights=[40, 40, 20])[0]
        chosen_genes = random.sample(AMR_GENES, min(n_genes, len(AMR_GENES)))
        for (gene, el_type, cls, sub, identity, coverage) in chosen_genes:
            amr_insert.append((
                str(uuid.uuid4()), isolate_id,
                gene, el_type, cls, sub,
                round(identity  + random.uniform(-2, 2), 2),
                round(coverage  + random.uniform(-5, 5), 2),
            ))

    print(f"Inserting {len(amr_insert)} AMR sequence records...")
    execute_values(cur, """
        INSERT INTO amr_sequences
          (sequence_id, isolate_id, gene_symbol, element_type,
           resistance_class, resistance_subclass,
           identity_percentage, coverage_percentage)
        VALUES %s
    """, amr_insert)
    conn.commit()

    # ── 5. WGS METRICS ──────────────────────────────────────────────────────
    wgs_insert = []
    for (isolate_id, org_name) in isolate_ids:
        tpl = random.choice(WGS_TEMPLATES)
        status, genotype, phenotype, sir, plasmid, genome_len, n50 = tpl
        wgs_insert.append((
            str(uuid.uuid4()), isolate_id,
            status, phenotype, genotype, plasmid,
            genome_len + random.randint(-200000, 200000),
            max(1000, n50 + random.randint(-50000, 50000)),
            sir,
        ))

    print(f"Inserting {len(wgs_insert)} WGS metric records...")
    execute_values(cur, """
        INSERT INTO wgs_metrics
          (wgs_id, isolate_id, quality_status, predicted_phenotype,
           genotype, plasmid, genome_length, n50_value, predicted_sir_profile)
        VALUES %s
    """, wgs_insert)
    conn.commit()

    # ── Summary ──────────────────────────────────────────────────────────────
    cur.execute("SELECT COUNT(*) FROM sites")
    print(f"\n✅ Seeding complete:")
    print(f"   Sites:          {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM water_samples")
    print(f"   Water samples:  {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM isolates")
    print(f"   Isolates:       {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM amr_sequences")
    print(f"   AMR sequences:  {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM wgs_metrics")
    print(f"   WGS metrics:    {cur.fetchone()[0]}")

    cur.close()
    conn.close()

if __name__ == "__main__":
    random.seed(42)   # reproducible
    seed()
